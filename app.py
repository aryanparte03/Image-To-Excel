import streamlit as st
from PIL import Image
import numpy as np
import pandas as pd
import easyocr
import io

import torch
from transformers import TableTransformerForObjectDetection, TableTransformerImageProcessor

# --- Model Config: Load Once ---
@st.cache_resource
def load_table_transformer():
    processor = TableTransformerImageProcessor(do_rescale=True, image_mean=[0.5], image_std=[0.5])
    model = TableTransformerForObjectDetection.from_pretrained(
        "microsoft/table-transformer-detection"
    )
    return processor, model

# Utility: Get EasyOCR Reader for Languages
@st.cache_resource
def get_easyocr_reader(langs):
    return easyocr.Reader(langs)

# --- UI Setup ---
st.set_page_config(page_title="Advanced Image-to-Excel Table Extraction")
st.title("Image-to-Excel Table Extraction App")

langs_all = ['en', 'hi', 'fr', 'de', 'ru', 'es', 'zh', 'ja', 'ko', 'ar']  # Add as many supported by EasyOCR
ocr_langs = st.multiselect("Select OCR language(s) for table extraction:", langs_all, default=['en'])

uploaded_files = st.file_uploader(
    "Upload one or more table images (.png, .jpg, .jpeg, .bmp, .tiff):",
    type=['png', 'jpg', 'jpeg', 'bmp', 'tiff'],
    accept_multiple_files=True
)

if not uploaded_files:
    st.info("Please upload table images to proceed.")
    st.stop()

reader = get_easyocr_reader(ocr_langs)
processor, model = load_table_transformer()

# Helper: Detect and parse table cells with Table Transformer
def detect_cells(image):
    inputs = processor(images=image, return_tensors="pt")
    with torch.no_grad():
        predictions = model(**inputs)
    target_sizes = torch.tensor([image.size[::-1]])
    results = processor.post_process_object_detection(
        predictions, threshold=0.8, target_sizes=target_sizes
    )[0]
    boxes = results['boxes']
    # Each box: (xmin, ymin, xmax, ymax)
    if len(boxes) < 1:
        return []
    # Sort by top-to-bottom, then left-to-right
    boxes = sorted(
        [tuple(map(int, box)) for box in boxes],
        key=lambda b: (b[1], b[0])
    )
    return boxes

# Helper: Convert detected cells into a table for OCR
def cells_to_table(image, cell_boxes):
    # Heuristic: group boxes into rows based on Y overlap
    tolerance = 15
    rows = []
    for box in cell_boxes:
        x1, y1, x2, y2 = box
        center_y = (y1 + y2) // 2
        placed = False
        for row in rows:
            if abs(row[0][1] + row[0][3]//2 - center_y) < tolerance:
                row.append(box)
                placed = True
                break
        if not placed:
            rows.append([box])
    # Sort cells in each row left-to-right
    rows = [sorted(r, key=lambda b: b[0]) for r in rows]
    # OCR: For each cell
    out = []
    for row in rows:
        row_text = []
        for x1, y1, x2, y2 in row:
            cell = image.crop((x1, y1, x2, y2))
            ocr = reader.readtext(np.array(cell), detail=0, paragraph=True)
            txt = " ".join(ocr).strip()
            row_text.append(txt)
        out.append(row_text)
    # Rectangular DataFrame
    max_cols = max(len(r) for r in out)
    normed = [r + [""]*(max_cols-len(r)) for r in out]
    return pd.DataFrame(normed)

# Helper: Convert DataFrame(s) to Excel file
def create_excel(sheets):
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for name, df in sheets:
        ws = wb.create_sheet(title=name[:28])
        for r_idx, row in df.iterrows():
            for c_idx, val in enumerate(row):
                ws.cell(row=r_idx+1, column=c_idx+1, value=val)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# --- Main Application Flow ---
dataframes = []
for idx, uploaded_file in enumerate(uploaded_files):
    st.write(f"### Table {idx+1}: {uploaded_file.name}")
    img = Image.open(uploaded_file).convert('RGB')
    st.image(img, caption=uploaded_file.name)
    with st.spinner("Analyzing and extracting table..."):
        boxes = detect_cells(img)
        if not boxes:
            st.warning("No table detected. Try with a clearer screenshot.")
            df = pd.DataFrame([["No table detected"]])
        else:
            df = cells_to_table(img, boxes)
    # User correction interface
    st.write("Edit detected table as needed before export:")
    df_edited = st.data_editor(df, num_rows="dynamic", key=f"editor_{idx}")
    dataframes.append((uploaded_file.name.split('.')[0], df_edited))

# --- Download Link ---
with st.expander("Download your Excel file"):
    excel_bytes = create_excel(dataframes)
    btn_label = "Download Excel"
    st.download_button(
        label=btn_label,
        data=excel_bytes,
        file_name="tables.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.success("Done! You can now download the merged Excel workbook for all your tables.")

st.caption("Built with Streamlit, Table Transformer, and EasyOCR.")
